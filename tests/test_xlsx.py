import os

from openpyxl import load_workbook

from conftest import RESOURCES_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь

def test_xlsx():
    xlsx_path = os.path.join(RESOURCES_PATH, 'file_example_XLSX_50.xlsx')
    workbook = load_workbook(xlsx_path)
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert 'Mara' == sheet.cell(row=3, column=2).value